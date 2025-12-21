# contracts_app/services.py
from datetime import datetime

from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from django.core.paginator import Paginator

from .models import Contract, Implementator, Work


def get_filtered_contracts(get_params):
    """Возвращает отфильтрованный QuerySet договоров по GET-параметрам"""
    contracts = (
        Contract.objects.select_related("implementator")
        .prefetch_related("aks", "works")
        .annotate(ak_count=Count("aks"))
        .order_by("-id")
    )

    if get_params:
        if q := get_params.get("q"):
            contracts = contracts.filter(
                Q(customer_name__icontains=q) | Q(customer_inn__icontains=q)
            )

        if ak_number := get_params.get("ak_number"):
            contracts = contracts.filter(aks__number=ak_number)

        if ak_district := get_params.get("ak_district"):
            contracts = contracts.filter(aks__district__name__icontains=ak_district.strip())

        if ak_address := get_params.get("ak_address"):
            contracts = contracts.filter(aks__address__icontains=ak_address.strip())

        if status := get_params.get("status"):
            contracts = contracts.filter(status=status)

        if impl := get_params.get("implementator"):
            contracts = contracts.filter(implementator_id=impl)

        if works := get_params.getlist("works"):
            contracts = contracts.filter(works__pk__in=works)

        checklist_filters = Q()
        if get_params.get("gos_services"):
            checklist_filters |= Q(gos_services=True)
        if get_params.get("oko"):
            checklist_filters |= Q(oko=True)
        if get_params.get("spolokh"):
            checklist_filters |= Q(spolokh=True)
        if checklist_filters:
            contracts = contracts.filter(checklist_filters)

        # Стадия подписания (хотя бы один включён)
        stage_filters = Q()
        if get_params.get("stage_to_be_signed"):
            stage_filters |= Q(contract_to_be_signed=True)
            if get_params.get("stage_signed"):
                stage_filters |= Q(contract_signed=True)
            if get_params.get("stage_trading_platform"):
                stage_filters |= Q(contract_signed_in_trading_platform=True)
            if get_params.get("stage_edo"):
                stage_filters |= Q(contract_signed_in_EDO=True)
            if get_params.get("stage_original_received"):
                stage_filters |= Q(contract_original_received=True)
            if get_params.get("stage_termination"):
                stage_filters |= Q(contract_termination=True)
            if stage_filters:
                contracts = contracts.filter(stage_filters)

        contracts = contracts.distinct()

    return contracts


def get_contract_list_context(request):
    """Возвращает контекст для contract_list (отфильтрованные договоры + справочники)"""
    get_params = request.GET

    contracts = get_filtered_contracts(get_params)

    # Сортировка
    ordering = get_params.get("order", "-created_at")
    if ordering == "created_at":
        contracts = contracts.order_by("created_at")
    elif ordering == "customer_name":
        contracts = contracts.order_by("customer_name")
    elif ordering == "-customer_name":
        contracts = contracts.order_by("-customer_name")
    else:
        contracts = contracts.order_by("-created_at")

    # Счётчики
    total_contracts = contracts.count()
    total_aks = contracts.aggregate(total=Sum("ak_count"))["total"] or 0

    # Пагинация
    paginator = Paginator(contracts, 10)
    page_number = get_params.get("page")
    page_obj = paginator.get_page(page_number)

    return {
        "page_obj": page_obj,
        "contracts": page_obj,
        "implementators": Implementator.objects.all().order_by("name"),
        "all_works": Work.objects.all().order_by("name"),
        "is_paginated": page_obj.has_other_pages(),
        "total_contracts": total_contracts,
        "total_aks": total_aks,
        "current_ordering": ordering,
    }


def update_contract_signing_stage(user, pk, stage):
    """Обновляет стадию подписания договора. Возвращает JsonResponse."""
    contract = get_object_or_404(Contract, pk=pk)

    if not (user.is_superuser or user.groups.filter(name="Administrators").exists()):
        return JsonResponse({"success": False, "message": "Нет прав"})

    # Сбрасываем все
    contract.contract_to_be_signed = False
    contract.contract_signed = False
    contract.contract_signed_in_trading_platform = False
    contract.contract_signed_in_EDO = False
    contract.contract_original_received = False
    contract.contract_termination = False

    # Включаем выбранный
    if stage == "to_be_signed":
        contract.contract_to_be_signed = True
    elif stage == "signed":
        contract.contract_signed = True
    elif stage == "trading_platform":
        contract.contract_signed_in_trading_platform = True
    elif stage == "edo":
        contract.contract_signed_in_EDO = True
    elif stage == "original_received":
        contract.contract_original_received = True
    elif stage == "termination":
        contract.contract_termination = True

    contract.save()

    return JsonResponse({"success": True, "message": "Стадия подписания обновлена"})


def export_contracts_to_excel(get_params):
    """Генерирует и возвращает HttpResponse с Excel-файлом"""
    contracts = get_filtered_contracts(get_params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Долгосрочные договоры"

    headers = [
        "№",
        "Заказчик",
        "ИНН",
        "Исполнитель",
        "Срок действия",
        "Статус",
        "Кол-во АК",
        "АК (номера)",
        "Работы",
        "Сумма общая",
        "В мес",
        "Госуслуги",
        "ОКО",
        "Сполох",
    ]
    ws.append(headers)

    # Стили заголовков
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center

    # Данные
    for idx, contract in enumerate(contracts, start=1):
        ak_numbers = (
            " | ".join(str(ak.number) for ak in contract.aks.all()) if contract.aks.exists() else ""
        )
        works_list = (
            " | ".join(w.name for w in contract.works.all()) if contract.works.exists() else ""
        )

        row = [
            idx,
            contract.customer_name or "",
            contract.customer_inn or "",
            str(contract.implementator),
            f"{contract.start_date} — {contract.end_date}",
            contract.get_status_display(),
            contract.aks.count(),
            ak_numbers,
            works_list,
            contract.total_amount or "",
            contract.monthly_amount or "",
            "Да" if contract.gos_services else "",
            "Да" if contract.oko else "",
            "Да" if contract.spolokh else "",
        ]
        ws.append(row)

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value or "")) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Ответ
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"договоры_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
