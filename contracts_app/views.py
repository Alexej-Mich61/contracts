# contracts_app/views.py
from datetime import datetime

from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from .forms import AKFormSet, ContractForm
from .models import Contract, Implementator, Work
from .permissions import ViewOnlyPermissionMixin


def permission_denied_view(request, exception=None):
    return render(request, "403.html", status=403)


def contract_list(request):
    # Базовый запрос
    contracts = (
        Contract.objects.select_related("implementator")
        .prefetch_related("aks", "works")
        .annotate(ak_count=Count("aks"))
    )
    implementators = Implementator.objects.all().order_by("name")
    all_works = Work.objects.all().order_by("name")

    # === ФИЛЬТРАЦИЯ ===
    if request.GET:
        # 1. Поиск по заказчику / ИНН — 100% регистронезависимый
        if q := request.GET.get("q"):
            contracts = contracts.filter(
                Q(customer_name__icontains=q) | Q(customer_inn__icontains=q)
            )

        # 2. Номер АК
        if ak_number := request.GET.get("ak_number"):
            contracts = contracts.filter(aks__number=ak_number)

        # 3. Район АК — регистронезависимый
        if ak_district := request.GET.get("ak_district"):
            contracts = contracts.filter(aks__district__name__icontains=ak_district.strip())

        # 4. Адрес АК — регистронезависимый
        if ak_address := request.GET.get("ak_address"):
            contracts = contracts.filter(aks__address__icontains=ak_address.strip())

        # 5. Статус
        if status := request.GET.get("status"):
            contracts = contracts.filter(status=status)

        # 6. Исполнитель
        if impl := request.GET.get("implementator"):
            contracts = contracts.filter(implementator_id=impl)

        # 7. Работы (множественный)
        if works := request.GET.getlist("works"):
            contracts = contracts.filter(works__pk__in=works)

        # 8. Чек-лист (хотя бы один включён)
        checklist_filters = Q()
        if request.GET.get("gos_services"):
            checklist_filters |= Q(gos_services=True)
        if request.GET.get("oko"):
            checklist_filters |= Q(oko=True)
        if request.GET.get("spolokh"):
            checklist_filters |= Q(spolokh=True)
        if checklist_filters:
            contracts = contracts.filter(checklist_filters)

        # Убираем дубли из-за join
        contracts = contracts.distinct()

    # ←←← НОВАЯ СОРТИРОВКА ←←←
    ordering = request.GET.get("order", "-created_at")  # по умолчанию: новые сверху

    if ordering == "created_at":
        contracts = contracts.order_by("created_at")  # от старых к новым
    elif ordering == "customer_name":
        contracts = contracts.order_by("customer_name")  # по алфавиту А → Я
    elif ordering == "-customer_name":
        contracts = contracts.order_by("-customer_name")  # по алфавиту Я → А
    else:
        contracts = contracts.order_by("-created_at")  # новые → старые (по умолчанию)

    # Счетчики
    total_contracts = contracts.count()
    total_aks = contracts.aggregate(total=Sum("ak_count"))["total"] or 0

    # Пагинация
    paginator = Paginator(contracts, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "contracts": page_obj,
        "implementators": implementators,
        "all_works": all_works,
        "is_paginated": page_obj.has_other_pages(),
        # Два счетчика
        "total_contracts": total_contracts,
        "total_aks": total_aks,
        "current_ordering": ordering,  # ← передаём текущую сортировку
    }
    return render(request, "contracts/contract_list.html", context)


def export_contracts_excel(request):
    contracts = (
        Contract.objects.select_related("implementator")
        .prefetch_related("aks", "works")
        .annotate(ak_count=Count("aks"))
        .order_by("-id")
    )

    if request.GET:
        if q := request.GET.get("q"):
            contracts = contracts.filter(
                Q(customer_name__icontains=q) | Q(customer_inn__icontains=q)
            )
        if ak_number := request.GET.get("ak_number"):
            contracts = contracts.filter(aks__number=ak_number)
        if ak_district := request.GET.get("ak_district"):
            contracts = contracts.filter(aks__district__name__icontains=ak_district.strip())
        if ak_address := request.GET.get("ak_address"):
            contracts = contracts.filter(aks__address__icontains=ak_address.strip())
        if status := request.GET.get("status"):
            contracts = contracts.filter(status=status)
        if impl := request.GET.get("implementator"):
            contracts = contracts.filter(implementator_id=impl)
        if works := request.GET.getlist("works"):
            contracts = contracts.filter(works__pk__in=works)
        checklist_filters = Q()
        if request.GET.get("gos_services"):
            checklist_filters |= Q(gos_services=True)
        if request.GET.get("oko"):
            checklist_filters |= Q(oko=True)
        if request.GET.get("spolokh"):
            checklist_filters |= Q(spolokh=True)
        if checklist_filters:
            contracts = contracts.filter(checklist_filters)

    contracts = contracts.distinct()

    wb = Workbook()
    ws = wb.active
    ws.title = "Долгосрочные договоры"

    headers = [
        "№",
        "Заказчик",
        "ИНН",
        "Исполнитель",
        "Срок действия",
        "Статус",
        "Кол-во АК",
        "АК (номера)",
        "Работы",
        "Сумма общая",
        "В мес",
        "Госуслуги",
        "ОКО",
        "Сполох",
    ]
    ws.append(headers)

    # Стили заголовков
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center

    # Данные
    for idx, contract in enumerate(contracts, start=1):
        # ←←← ИСПРАВЛЕНО: str() для всех значений ←←←
        ak_numbers = (
            " | ".join(str(ak.number) for ak in contract.aks.all()) if contract.aks.exists() else ""
        )
        works_list = (
            " | ".join(w.name for w in contract.works.all()) if contract.works.exists() else ""
        )

        row = [
            idx,
            contract.customer_name or "",
            contract.customer_inn or "",
            str(contract.implementator),
            f"{contract.start_date} — {contract.end_date}",
            contract.get_status_display(),
            contract.aks.count(),
            ak_numbers,
            works_list,
            contract.total_amount or "",
            contract.monthly_amount or "",
            "Да" if contract.gos_services else "",
            "Да" if contract.oko else "",
            "Да" if contract.spolokh else "",
        ]
        ws.append(row)

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value or "")) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Ответ
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"договоры_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class ContractDetailView(DetailView):
    model = Contract
    template_name = "contracts/contract_detail.html"
    context_object_name = "contract"


class ContractCreateView(ViewOnlyPermissionMixin, CreateView):
    model = Contract
    form_class = ContractForm
    template_name = "contracts/contract_form.html"
    success_url = reverse_lazy("contracts:contract_list")

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data["ak_formset"] = AKFormSet(self.request.POST, self.request.FILES)
        else:
            data["ak_formset"] = AKFormSet()
        return data

    def form_valid(self, form):
        response = super().form_valid(form)
        # Сохраняем АК формсет
        ak_formset = self.get_context_data()["ak_formset"]
        ak_formset.instance = self.object
        if ak_formset.is_valid():
            ak_formset.save()

        # Добавляем сообщение в сессию
        messages.success(self.request, "Договор успешно создан!")
        return response

    def form_invalid(self, form):
        # Добавляем формсет в контекст, чтобы ошибки отобразились
        context = self.get_context_data()
        context["ak_formset"] = AKFormSet(self.request.POST, self.request.FILES)
        return self.render_to_response(context)


class ContractUpdateView(ViewOnlyPermissionMixin, UpdateView):
    model = Contract
    form_class = ContractForm
    template_name = "contracts/contract_form.html"
    success_url = reverse_lazy("contracts:contract_list")

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data["ak_formset"] = AKFormSet(
                self.request.POST, self.request.FILES, instance=self.object
            )
        else:
            data["ak_formset"] = AKFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        response = super().form_valid(form)
        ak_formset = self.get_context_data()["ak_formset"]
        ak_formset.instance = self.object
        if ak_formset.is_valid():
            ak_formset.save()

        messages.success(self.request, "Договор успешно обновлён!")
        return response

    def form_invalid(self, form):
        context = self.get_context_data()
        context["ak_formset"] = AKFormSet(
            self.request.POST, self.request.FILES, instance=self.object
        )
        return self.render_to_response(context)


class ContractDeleteView(ViewOnlyPermissionMixin, DeleteView):
    model = Contract
    success_url = reverse_lazy("contracts:contract_list")

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Договор успешно удалён.")
        return super().delete(request, *args, **kwargs)


@require_POST
def update_checklist(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    contract.gos_services = "gos_services" in request.POST
    contract.oko = "oko" in request.POST
    contract.spolokh = "spolokh" in request.POST
    contract.save()
    return JsonResponse({"success": True, "message": "Чек-лист обновлён"})
